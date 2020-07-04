#-*- coding: utf-8-*-
#@Time : 2020/6/15 22:00
#@Author : beibei
#@Email : libei0106@qq.com
#@File : class_lianxi.py
#执行文件
#调用
from zuoye_yongli_zhixing.http_DuQv import read_data
from zuoye_yongli_zhixing.http_giuzong import http_request
from openpyxl import load_workbook
#获取所有测试用例数据
Token=None#全局变量，初始值设置为None
def run():
    global Token#在这里申明，函数外的Token 和函数内的Token是同一个值
    all_case=read_data('login.xlsx','cz')
    print('获取到的所有测试数据是:',all_case)
    for i in range(len(all_case)):#在http——response进行请求的时候，判断是否是登录请求
        test_data = all_case[i]
        # if test_data[0]==1；
        # if test_data[1]=='登录'：判断两边是否相等  比较运算符
        ip="http://120.78.128.25:8766"
        response = http_request(ip + test_data[4], eval(test_data[5]), token=Token,
                                method=test_data[3])
        if 'login' in test_data[4]:#成员运算符
            #他就是一个登录的用例
            Token="Bearer "+response['data']['token_info']['token']
        expected = eval(test_data[6])  # 期望值
        print("最后的结果:",response)

        #开始写入结果
        wb=load_workbook('login.xlsx')
        sheet=wb['cz']
        #定位单元格存值、行、列、值
        sheet.cell(row=test_data[0]+1,column=8).value=str(response)
        test_data[6]
        #保存
        wb.save('login.xlsx')
#调用函数
run()