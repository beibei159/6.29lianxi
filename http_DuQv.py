#-*- coding: utf-8-*-
#@Time : 2020/6/15 22:00
#@Author : beibei
#@Email : libei0106@qq.com
#@File : class_lianxi.py
from openpyxl import load_workbook
def read_data(file_name,sheet_name):#
        wb=load_workbook(file_name)
        chongzhi=wb[sheet_name]
        LL=[]#存储所有数据
        for j in range(2,chongzhi.max_row+1):
            kk=[]#存储某一行的数据
            for i in range(0,chongzhi.max_column-2):
                kk.append(chongzhi.cell(row=j,column=i+1).value)
                # print(chongzhi.cell(row=2,column=i+1).value)
            # print(kk)
            LL.append(kk)
        # print(LL)
        return LL#返回所有测试用例数据
if __name__ == '__main__':

    LL=read_data('login.xlsx','cz')
    print("所有测试数据为:",LL)
